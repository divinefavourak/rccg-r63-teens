from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill
from .models import Ticket, BulkUpload, TicketAuditLog, CheckInRecord


@admin.register(Ticket)
class TicketAdmin(admin.ModelAdmin):
    """Admin configuration for Ticket model"""
    list_display = (
        'ticket_id', 'full_name', 'age', 'category', 'gender',
        'province', 'status', 'payment_status', 'proof_link', 'registered_at'
    )
    list_filter = (
        'status', 'payment_status', 'category', 'gender', 'province',
        'zone', 'area', 'registered_at'
    )
    search_fields = (
        'ticket_id', 'full_name', 'email', 'phone',
        'province', 'zone', 'area', 'parish'
    )
    readonly_fields = ('ticket_id', 'registered_at', 'created_at', 'updated_at', 'qr_code')
    date_hierarchy = 'registered_at'
    list_per_page = 50
    
    actions = ['export_to_excel', 'approve_tickets', 'reject_tickets']
    
    fieldsets = (
        ('Ticket Information', {
            'fields': ('ticket_id', 'status', 'notes', 'qr_code')
        }),
        ('Payment Information', {
            'fields': ('payment_status', 'proof_of_payment')
        }),
        ('Personal Information', {
            'fields': ('full_name', 'age', 'date_of_birth', 'category', 'gender')
        }),
        ('Contact Information', {
            'fields': ('phone', 'email')
        }),
        ('Church Hierarchy', {
            'fields': ('province', 'zone', 'area', 'parish', 'department')
        }),
        ('Medical Information', {
            'fields': ('medical_conditions', 'medications', 'dietary_restrictions'),
            'classes': ('collapse',)
        }),
        ('Emergency Contact', {
            'fields': ('emergency_contact', 'emergency_phone', 'emergency_relationship')
        }),
        ('Parent/Guardian Information', {
            'fields': ('parent_name', 'parent_email', 'parent_phone', 'parent_relationship')
        }),
        ('Registration Details', {
            'fields': ('registered_at', 'registered_by', 'approved_at', 'approved_by')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    def proof_link(self, obj):
        if obj.proof_of_payment:
            from django.utils.html import format_html
            return format_html('<a href="{}" target="_blank">View Proof</a>', obj.proof_of_payment.url)
        return "-"
    proof_link.short_description = "Proof of Payment"
    
    def export_to_excel(self, request, queryset):
        """Export selected tickets to Excel"""
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registrations"
        
        # Define headers
        headers = [
            'Ticket ID', 'Full Name', 'Age', 'Category', 'Gender',
            'Province', 'Zone', 'Area', 'Parish', 'Department',
            'Phone', 'Email', 'Status', 'Registered At',
            'Emergency Contact', 'Emergency Phone', 'Emergency Relationship',
            'Parent Name', 'Parent Email', 'Parent Phone',
            'Medical Conditions', 'Medications', 'Dietary Restrictions'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Add data
        for row_num, ticket in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=ticket.ticket_id)
            ws.cell(row=row_num, column=2, value=ticket.full_name)
            ws.cell(row=row_num, column=3, value=ticket.age)
            ws.cell(row=row_num, column=4, value=ticket.get_category_display())
            ws.cell(row=row_num, column=5, value=ticket.get_gender_display())
            ws.cell(row=row_num, column=6, value=ticket.get_province_display())
            ws.cell(row=row_num, column=7, value=ticket.zone)
            ws.cell(row=row_num, column=8, value=ticket.area)
            ws.cell(row=row_num, column=9, value=ticket.parish)
            ws.cell(row=row_num, column=10, value=ticket.department)
            ws.cell(row=row_num, column=11, value=ticket.phone)
            ws.cell(row=row_num, column=12, value=ticket.email or '')
            ws.cell(row=row_num, column=13, value=ticket.get_status_display())
            ws.cell(row=row_num, column=14, value=ticket.registered_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=15, value=ticket.emergency_contact)
            ws.cell(row=row_num, column=16, value=ticket.emergency_phone)
            ws.cell(row=row_num, column=17, value=ticket.emergency_relationship)
            ws.cell(row=row_num, column=18, value=ticket.parent_name)
            ws.cell(row=row_num, column=19, value=ticket.parent_email)
            ws.cell(row=row_num, column=20, value=ticket.parent_phone)
            ws.cell(row=row_num, column=21, value=ticket.medical_conditions)
            ws.cell(row=row_num, column=22, value=ticket.medications)
            ws.cell(row=row_num, column=23, value=ticket.dietary_restrictions)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'registrations_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        wb.save(response)
        return response
    
    export_to_excel.short_description = "Export selected to Excel"
    
    def approve_tickets(self, request, queryset):
        """Bulk approve selected tickets"""
        updated = queryset.filter(status=Ticket.Status.PENDING).update(
            status=Ticket.Status.APPROVED,
            approved_at=timezone.now(),
            approved_by=request.user
        )
        self.message_user(request, f'{updated} ticket(s) approved successfully.')
    
    approve_tickets.short_description = "Approve selected tickets"
    
    def reject_tickets(self, request, queryset):
        """Bulk reject selected tickets"""
        updated = queryset.filter(status=Ticket.Status.PENDING).update(
            status=Ticket.Status.REJECTED
        )
        self.message_user(request, f'{updated} ticket(s) rejected.')
    
    reject_tickets.short_description = "Reject selected tickets"
    
    def get_queryset(self, request):
        """Filter tickets based on user's province if they're a coordinator"""
        qs = super().get_queryset(request)
        if request.user.is_superuser or request.user.is_admin:
            return qs
        if request.user.is_coordinator and request.user.province:
            return qs.filter(province=request.user.province)
        return qs.none()



@admin.register(BulkUpload)
class BulkUploadAdmin(admin.ModelAdmin):
    """Admin configuration for BulkUpload model"""
    list_display = ('filename', 'uploaded_by', 'status', 'total_records', 'created_at')
    list_filter = ('status', 'created_at')
    search_fields = ('filename', 'uploaded_by__username')
    readonly_fields = ('created_at', 'processed_at', 'error_log')


@admin.register(TicketAuditLog)
class TicketAuditLogAdmin(admin.ModelAdmin):
    """Admin configuration for TicketAuditLog model"""
    list_display = ('user', 'action', 'ticket', 'timestamp')
    list_filter = ('action', 'timestamp')
    search_fields = ('user__username', 'ticket__ticket_id', 'ticket__full_name')
    readonly_fields = ('timestamp',)
    date_hierarchy = 'timestamp'
    
@admin.register(CheckInRecord)
class CheckInRecordAdmin(admin.ModelAdmin):
    """Admin configuration for CheckInRecord model"""
    list_display = ('ticket', 'checked_in_by', 'checked_in_at', 'check_in_method')
    list_filter = ('check_in_method', 'checked_in_at')
    search_fields = ('ticket__ticket_id', 'ticket__full_name', 'checked_in_by__username')
    readonly_fields = ('checked_in_at',)
    date_hierarchy = 'checked_in_at'